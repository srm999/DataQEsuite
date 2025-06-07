import pandas as pd
import numpy as np
import re
import os
import gc
import traceback
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional, Union, Any
import core.custom_logger as custom_logger


log = custom_logger.customLogger()

# Shared utility functions
def normalize_key_value(value, logger=None):
    """Normalize a key value for consistent comparison"""
    try:
        if pd.isna(value):
            return "NA"
        value = str(value).lower()
        value = re.sub(r'[_\s]+', '', value).strip()
        return value
    except Exception as e:
        log.error(f"Error normalizing key value: {str(e)}")
        return str(value) if value is not None else "NA"

def compare_values(src_val, tgt_val):
    """Compare two values considering data types and NaN values"""
    # Handle NaN values
    if pd.isna(src_val) and pd.isna(tgt_val):
        return True
    elif pd.isna(src_val) or pd.isna(tgt_val):
        return False
    # Handle numeric values
    elif isinstance(src_val, (int, float)) and isinstance(tgt_val, (int, float)):
        return src_val == tgt_val
    # Handle string values (case insensitive)
    elif isinstance(src_val, str) and isinstance(tgt_val, str):
        return src_val.lower() == tgt_val.lower()
    # Any other type
    else:
        return src_val == tgt_val

def find_matching_columns(source_columns, target_columns, case_sensitive=False):
    """Find matching columns between source and target, handling case sensitivity"""
    if case_sensitive:
        return [col for col in source_columns if col in target_columns]
    
    src_col_map = {col.lower(): col for col in source_columns}
    tgt_col_map = {col.lower(): col for col in target_columns}
    
    shared_col_keys = set(src_col_map.keys()).intersection(set(tgt_col_map.keys()))
    src_shared_cols = [src_col_map[col_key] for col_key in shared_col_keys]
    tgt_shared_cols = [tgt_col_map[col_key] for col_key in shared_col_keys]
    
    return src_shared_cols, tgt_shared_cols

def find_key_columns(df, key_columns, logger=None):
    """Find actual key column names (case-insensitive)"""
    try:
        actual_key_columns = []
        
        for key_col in key_columns:
            # Direct match
            if key_col in df.columns:
                actual_key_columns.append(key_col)
                continue
                
            # Case-insensitive match
            matches = [col for col in df.columns if col.lower() == key_col.lower()]
            if matches:
                actual_key_columns.append(matches[0])
                continue
                
            # Match without underscores/spaces
            norm_key = re.sub(r'[_\s]+', '', key_col.lower())
            matches = [col for col in df.columns if re.sub(r'[_\s]+', '', col.lower()) == norm_key]
            if matches:
                actual_key_columns.append(matches[0])
                continue
            
            # No match found
            log.error(f"Error: Could not find column matching '{key_col}' in the dataframe")
        
        return actual_key_columns
    except Exception as e:
        log.error(f"Error finding key columns: {str(e)}")
        return []

def create_composite_keys(df, key_columns, logger=None):
    """Create normalized composite keys based on key columns"""
    try:
        def create_key(row):
            key_values = []
            for k in key_columns:
                val = row[k]
                key_values.append(normalize_key_value(val, logger))
            return tuple(key_values)
        
        # Check for null values in key columns
        for col in key_columns:
            null_count = df[col].isna().sum()
            if null_count > 0:
                log.info(f"Column '{col}' has {null_count} nulls")
        
        df['__key'] = df.apply(create_key, axis=1)
        return df
    except Exception as e:
        log.error(f"Error creating composite keys: {str(e)}")
        log.error(traceback.format_exc())
        return df


class DataComparator:
    """
    Handles efficient comparison between source and target datasets.
    Focuses on identifying differences and value mismatches.
    """
    
    def __init__(self, logger=None):
        self.log = custom_logger.customLogger()
        self.mismatch_details = {}
    
    def _normalize_column_names(self, src_df, tgt_df):
        """
        Create mappings between lowercase column names and actual column names.
        Identify shared columns between source and target dataframes.
        """
        # Create case-insensitive mappings
        src_col_map = {col.lower(): col for col in src_df.columns}
        tgt_col_map = {col.lower(): col for col in tgt_df.columns}
        
        # Find shared columns (case-insensitive)
        shared_col_keys = sorted(list(set(src_col_map.keys()).intersection(set(tgt_col_map.keys()))))
        
        if not shared_col_keys:
            self.log.error(f"DataFrames have no columns in common. Source Columns - {src_df.columns} and Target columns - {tgt_df.columns}")
            raise ValueError("DataFrames have no columns in common")
        
        # Map lowercase keys back to actual column names
        src_shared_cols = [src_col_map[col_key] for col_key in shared_col_keys]
        tgt_shared_cols = [tgt_col_map[col_key] for col_key in shared_col_keys]
        
        return {
            'src_col_map': src_col_map,
            'tgt_col_map': tgt_col_map,
            'shared_col_keys': shared_col_keys,
            'src_shared_cols': src_shared_cols,
            'tgt_shared_cols': tgt_shared_cols
        }
    
    def _validate_key_columns(self, key_columns, shared_col_info):
        """Validate and find key columns in both dataframes."""
        if not key_columns:
            return shared_col_info['src_shared_cols'], shared_col_info['tgt_shared_cols']
            
        shared_col_keys = shared_col_info['shared_col_keys']
        src_shared_cols = shared_col_info['src_shared_cols']
        tgt_shared_cols = shared_col_info['tgt_shared_cols']
        
        # Create sets for faster lookups
        key_cols_lower = [col.lower() for col in key_columns]
        
        valid_key_indices = []
        
        # Try exact case-insensitive matching
        for i, col_key in enumerate(shared_col_keys):
            if col_key in key_cols_lower:
                valid_key_indices.append(i)
        
        # If not all keys found, try alphanumeric matching (ignoring special chars)
        if len(valid_key_indices) < len(key_columns):
            key_cols_alphanumeric = [''.join(c for c in col.lower() if c.isalnum()) for col in key_columns]
            shared_col_alphanumeric = [''.join(c for c in col.lower() if c.isalnum()) for col in shared_col_keys]
            
            for i, col_alphanum in enumerate(shared_col_alphanumeric):
                if col_alphanum in key_cols_alphanumeric and i not in valid_key_indices:
                    valid_key_indices.append(i)
        
        if not valid_key_indices:
            raise ValueError("Key columns not found in both dataframes")
            
        # Get actual column names for the valid keys
        src_keys = [src_shared_cols[i] for i in valid_key_indices]
        tgt_keys = [tgt_shared_cols[i] for i in valid_key_indices]
        
        return src_keys, tgt_keys
    
    def compare_dataframes(self, src_df, tgt_df, key_columns=None, chunk_size=500000):
        """
        Compare large dataframes in chunks to minimize memory usage.
        Works with both row hash comparison and key-based comparison.
        
        Args:
            src_df, tgt_df: DataFrames to compare
            key_columns: Columns that uniquely identify records (if None, uses row hash comparison)
            chunk_size: Size of chunks to process at once
            
        Returns:
            tuple: (diff_dataframe, summary_dataframe)
        """
        try:
            # Calculate number of chunks based on total rows
            total_rows = max(len(src_df), len(tgt_df))
            chunks = (total_rows + chunk_size - 1) // chunk_size
            
            # Validate key columns if provided
            valid_src_keys = None
            if key_columns:
                self.log.info(f"Using key-based comparison with keys: {key_columns}")
                shared_col_info = self._normalize_column_names(src_df, tgt_df)
                valid_src_keys, valid_tgt_keys = self._validate_key_columns(key_columns, shared_col_info)
                if not valid_src_keys:
                    self.log.warning("Key validation failed, falling back to row hash comparison")
            
            # Initialize result containers
            src_diff_chunks = []
            tgt_diff_chunks = []
            value_mismatch_chunks = []
            total_diff_count = 0
            value_mismatch_count = 0
            self.mismatch_details = {}

            self.log.info(f"Comparing {len(src_df)} source rows with {len(tgt_df)} target rows in {chunks} chunks")
            
            for i in range(chunks):
                start_idx = i * chunk_size
                end_idx = min((i + 1) * chunk_size, total_rows)
                
                # Extract chunks
                src_df_chunk = src_df.iloc[start_idx:end_idx] if start_idx < len(src_df) else src_df.iloc[0:0]
                tgt_df_chunk = tgt_df.iloc[start_idx:end_idx] if start_idx < len(tgt_df) else tgt_df.iloc[0:0]

                self.log.info(f"Processing chunk {i+1}/{chunks}: rows {start_idx} - {end_idx}")
                
                # Compare chunks using appropriate method
                if valid_src_keys:
                    # Key-based comparison
                    chunk_result = self._compare_with_keys(src_df_chunk, tgt_df_chunk, valid_src_keys)
                    
                    # Store results
                    if len(chunk_result['in_src_df_not_in_tgt_df']) > 0:
                        src_diff_chunks.append(chunk_result['in_src_df_not_in_tgt_df'])
                    
                    if len(chunk_result['in_tgt_df_not_in_src_df']) > 0:
                        tgt_diff_chunks.append(chunk_result['in_tgt_df_not_in_src_df'])
                    
                    if 'value_mismatch_df' in chunk_result and len(chunk_result['value_mismatch_df']) > 0:
                        value_mismatch_chunks.append(chunk_result['value_mismatch_df'])
                        value_mismatch_count += chunk_result['value_mismatches']
                    
                    # Merge mismatch details
                    if 'mismatch_details' in chunk_result:
                        self.mismatch_details.update(chunk_result['mismatch_details'])
                        
                    total_diff_count += chunk_result['total_diff_count']
                else:
                    # Row hash comparison
                    chunk_result = self.df_hash_for_compare(src_df_chunk, tgt_df_chunk)
                    
                    # Store results
                    if len(chunk_result['in_src_df_not_in_tgt_df']) > 0:
                        src_diff_chunks.append(chunk_result['in_src_df_not_in_tgt_df'])
                    
                    if len(chunk_result['in_tgt_df_not_in_src_df']) > 0:
                        tgt_diff_chunks.append(chunk_result['in_tgt_df_not_in_src_df'])
                        
                    total_diff_count += chunk_result['total_diff_count']

                if (i+1) % 5 == 0 or i+1 == chunks:
                    self.log.info(f"Processed {i+1}/{chunks} Chunk. Current diff count: {total_diff_count}")
                    if value_mismatch_count > 0:
                        self.log.info(f"Current value mismatch count: {value_mismatch_count}")
            
            # Combine all results
            in_src_not_in_tgt = pd.concat(src_diff_chunks) if src_diff_chunks else pd.DataFrame(columns=list(src_df.columns) + ['__source', '__mismatch_type'])
            in_tgt_not_in_src = pd.concat(tgt_diff_chunks) if tgt_diff_chunks else pd.DataFrame(columns=list(tgt_df.columns) + ['__source', '__mismatch_type'])
            value_mismatches = pd.concat(value_mismatch_chunks) if value_mismatch_chunks else pd.DataFrame()

            # Log summary of results before combining
            #self.log.info(f"Comparison results summary:")
            #self.log.info(f"  Rows only in source: {len(in_src_not_in_tgt)}")
            #self.log.info(f"  Rows only in target: {len(in_tgt_not_in_src)}")
            #self.log.info(f"  Value mismatches: {value_mismatch_count}")
            #self.log.info(f"  Total diff count: {total_diff_count}")

            # Combine all differences
            all_diffs = []
            if not in_src_not_in_tgt.empty:
                all_diffs.append(in_src_not_in_tgt)
            if not in_tgt_not_in_src.empty:
                all_diffs.append(in_tgt_not_in_src)
            if not value_mismatches.empty:
                all_diffs.append(value_mismatches)
                
            df_cmp_result_data = pd.concat(all_diffs) if all_diffs else pd.DataFrame()
            is_identical = total_diff_count == 0

            # Double-check value mismatch count from the result dataframe
            actual_value_mismatches = 0
            if not df_cmp_result_data.empty and '__mismatch_type' in df_cmp_result_data.columns:
                mismatch_counts = df_cmp_result_data['__mismatch_type'].value_counts().to_dict()
                if 'Value-Mismatch' in mismatch_counts:
                    # Value mismatches are counted as 2 rows (source and target)
                    actual_value_mismatches = mismatch_counts['Value-Mismatch'] // 2
                    #self.log.info(f"Actual value mismatches detected in result: {actual_value_mismatches}")
                    
                    # Use the higher count
                    value_mismatch_count = max(value_mismatch_count, actual_value_mismatches)

            # Create summary Dataframe - this is critical for correct counts
            summary_data = {
                'Metric': ['Total Source Rows', 'Total Target Rows', 'Rows Only in Source', 
                        'Rows Only in Target', 'Value Mismatches', 'Total Differences', 'Comparison Result'],
                'Value': [len(src_df), len(tgt_df), len(in_src_not_in_tgt), len(in_tgt_not_in_src), 
                        value_mismatch_count, total_diff_count, "Identical" if is_identical else "Differences found"]
            }
            cmp_summary_df = pd.DataFrame(summary_data)
            
            # Log the final summary for debugging
            self.log.info(f"Final comparison summary: {dict(zip(summary_data['Metric'], summary_data['Value']))}")

            return df_cmp_result_data, cmp_summary_df
                    
        except Exception as e:
            self.log.error(f"Failed: Comparison of source and target data - {e}")
            self.log.error(traceback.format_exc())
            raise ValueError(f"Data comparison failed: {e}")
    
    def df_hash_for_compare(self, src_df, tgt_df):
        """
        Efficiently compare two dataframes using hashes of entire rows.
        This method treats each row as a whole and doesn't analyze individual fields.
        """
        # Create copies to avoid modifying the originals
        src_df = src_df.copy()
        tgt_df = tgt_df.copy()

        # Normalize column names and find shared columns
        shared_col_info = self._normalize_column_names(src_df, tgt_df)
            
        # Create hash values for each row
        # Use numpy arrays for better memory efficiency
        src_hashes = pd.util.hash_pandas_object(src_df[shared_col_info['src_shared_cols']], index=False).values
        tgt_hashes = pd.util.hash_pandas_object(tgt_df[shared_col_info['tgt_shared_cols']], index=False).values
        
        # Convert to sets for fast lookup
        src_hash_set = set(src_hashes)
        tgt_hash_set = set(tgt_hashes)
        
        # Find differences
        only_in_src = src_hash_set - tgt_hash_set
        only_in_tgt = tgt_hash_set - src_hash_set

        # Create boolean masks instead of filtering directly
        src_mask = np.isin(src_hashes, list(only_in_src))
        tgt_mask = np.isin(tgt_hashes, list(only_in_tgt))
        
        # Get the difference dataframes
        src_diff = src_df.loc[src_mask]
        tgt_diff = tgt_df.loc[tgt_mask]
        
        # Add a source identifier column to each dataframe
        if not src_diff.empty:
            src_diff = src_diff.copy()
            src_diff['__source'] = 'Source'
            src_diff['__mismatch_type'] = 'Source-only-row'
        
        if not tgt_diff.empty:
            tgt_diff = tgt_diff.copy()
            tgt_diff['__source'] = 'Target'
            tgt_diff['__mismatch_type'] = 'Target-only-row'
            
        # Create result dictionary with standardized keys
        result_dict = {
            'in_src_df_not_in_tgt_df': src_diff,
            'in_tgt_df_not_in_src_df': tgt_diff,
            'total_src_rows': len(src_df),
            'total_tgt_rows': len(tgt_df),
            'rows_only_in_src': len(src_diff),
            'rows_only_in_tgt': len(tgt_diff),
            'value_mismatches': 0,  # Basic hash compare doesn't detect value mismatches
            'total_diff_count': len(src_diff) + len(tgt_diff),
            'is_identical': len(src_diff) == 0 and len(tgt_diff) == 0,
            'mismatch_details': {}  # No mismatch details for basic hash compare
        }
        
        return result_dict
    
    def _compare_with_keys(self, src_df, tgt_df, key_columns):
        try:
            # 1. Extract just the key columns for comparison
            src_keys_df = src_df[key_columns].copy()
            tgt_keys_df = tgt_df[key_columns].copy()
                       
            # 2. Create normalized representation of keys
            # This ensures case-insensitive comparison and handles spaces/special chars
            src_df['__key_tuple'] = src_keys_df.apply(lambda row: tuple(normalize_key_value(row[col]) 
                                                        for col in key_columns), axis=1)
            tgt_df['__key_tuple'] = tgt_keys_df.apply(lambda row: tuple(normalize_key_value(row[col]) 
                                                        for col in key_columns), axis=1)
            
            # 3. Find unique keys in each dataset
            unique_src_keys = set(src_df['__key_tuple'])
            unique_tgt_keys = set(tgt_df['__key_tuple'])
                        
            # 4. Determine which keys exist only in source or only in target
            only_in_src_keys = unique_src_keys - unique_tgt_keys
            only_in_tgt_keys = unique_tgt_keys - unique_src_keys
            common_keys = unique_src_keys.intersection(unique_tgt_keys)
                        
            # 5. Create masks for rows that are only in source or only in target
            src_key_in_src_only = src_df['__key_tuple'].isin(only_in_src_keys)
            tgt_key_in_tgt_only = tgt_df['__key_tuple'].isin(only_in_tgt_keys)
                        
            # 6. Extract rows that are only in source or only in target
            in_src_not_in_tgt = src_df.loc[src_key_in_src_only].copy() if any(src_key_in_src_only) else pd.DataFrame(columns=src_df.columns)
            in_tgt_not_in_src = tgt_df.loc[tgt_key_in_tgt_only].copy() if any(tgt_key_in_tgt_only) else pd.DataFrame(columns=tgt_df.columns)
            
            # Add source identifier and mismatch type
            if not in_src_not_in_tgt.empty:
                in_src_not_in_tgt['__source'] = 'Source'
                in_src_not_in_tgt['__mismatch_type'] = 'Source-only-row'
            if not in_tgt_not_in_src.empty:
                in_tgt_not_in_src['__source'] = 'Target'
                in_tgt_not_in_src['__mismatch_type'] = 'Target-only-row'
            
                       
            # 7. Find rows with matching keys but different values - FULLY OPTIMIZED VERSION
            value_mismatches = []
            value_mismatch_details = {}
            
            # For cases with many rows per key, we need a different optimization approach
            if common_keys:
                # Processing progress tracking
                total_keys = len(common_keys)
                self.log.info(f"Processing {total_keys} common keys")
                keys_processed = 0
                last_progress = 0
                
                # Get shared non-key columns
                non_key_columns = [col for col in src_df.columns if col not in key_columns and col != '__key_tuple']
                
                # Process in batches to avoid memory issues
                batch_size = 50000  # Adjust based on your available memory
                key_batches = [list(common_keys)[i:i+batch_size] for i in range(0, len(common_keys), batch_size)]
                
                for batch in key_batches:
                    # Filter rows for this batch of keys
                    src_batch = src_df[src_df['__key_tuple'].isin(batch)]
                    tgt_batch = tgt_df[tgt_df['__key_tuple'].isin(batch)]
                    
                    # Group by key tuple for faster access
                    src_grouped = src_batch.groupby('__key_tuple')
                    tgt_grouped = tgt_batch.groupby('__key_tuple')
                    
                    # Process each key in the batch
                    for key in batch:
                        if key not in src_grouped.groups or key not in tgt_grouped.groups:
                            continue
                        
                        # Get all rows with this key
                        src_rows = src_grouped.get_group(key)
                        tgt_rows = tgt_grouped.get_group(key)
                        
                       
                        # Create a hashable representation of non-key values for each row
                        src_value_hashes = {}
                        for s_idx, src_row in src_rows.iterrows():
                            value_hash = tuple(
                                normalize_key_value(src_row[col]) if col in src_row else None
                                for col in non_key_columns
                            )
                            if value_hash not in src_value_hashes:
                                src_value_hashes[value_hash] = []
                            src_value_hashes[value_hash].append((s_idx, src_row))
                        
                        tgt_value_hashes = {}
                        for t_idx, tgt_row in tgt_rows.iterrows():
                            value_hash = tuple(
                                normalize_key_value(tgt_row[col]) if col in tgt_row else None
                                for col in non_key_columns
                            )
                            if value_hash not in tgt_value_hashes:
                                tgt_value_hashes[value_hash] = []
                            tgt_value_hashes[value_hash].append((t_idx, tgt_row))
                        
                        # Find hashes in source not present in target
                        src_only_hashes = set(src_value_hashes.keys()) - set(tgt_value_hashes.keys())
                        
                        # Find hashes in target not present in source
                        tgt_only_hashes = set(tgt_value_hashes.keys()) - set(src_value_hashes.keys())
                        
                        # For each source-only hash, report the corresponding rows as mismatches
                        for hash_val in src_only_hashes:
                            for s_idx, src_row in src_value_hashes[hash_val]:
                                # Add source row
                                src_dict = src_row.to_dict()
                                src_dict['__source'] = 'Source'
                                src_dict['__mismatch_type'] = 'Value-Mismatch'
                                value_mismatches.append(src_dict)
                                
                                # Find which columns differ
                                mismatched_cols = []
                                for hash_val_tgt in tgt_value_hashes:
                                    for i, (val_src, val_tgt) in enumerate(zip(hash_val, hash_val_tgt)):
                                        if val_src != val_tgt:
                                            mismatched_cols.append(non_key_columns[i])
                                
                                # Record mismatch details
                                if key not in value_mismatch_details:
                                    value_mismatch_details[key] = list(set(mismatched_cols))
                        
                        # For each target-only hash, report the corresponding rows as mismatches
                        for hash_val in tgt_only_hashes:
                            for t_idx, tgt_row in tgt_value_hashes[hash_val]:
                                # Add target row
                                tgt_dict = tgt_row.to_dict()
                                tgt_dict['__source'] = 'Target'
                                tgt_dict['__mismatch_type'] = 'Value-Mismatch'
                                value_mismatches.append(tgt_dict)
                        
                        # Track progress
                        keys_processed += 1
                        progress = int((keys_processed / total_keys) * 100)
                        if progress >= last_progress + 10:  # Log every 10% progress
                            self.log.info(f"Progress: {progress}% ({keys_processed}/{total_keys} keys processed)")
                            last_progress = progress
                
                       
            # 8. Create dataframe of value mismatches
            value_mismatch_df = pd.DataFrame(value_mismatches) if value_mismatches else pd.DataFrame(columns=src_df.columns)
            
                       
            # 9. Combine all differences
            all_diffs = []
            if not in_src_not_in_tgt.empty:
                all_diffs.append(in_src_not_in_tgt)
            if not in_tgt_not_in_src.empty:
                all_diffs.append(in_tgt_not_in_src)
            if not value_mismatch_df.empty:
                all_diffs.append(value_mismatch_df)
            
            # Clean up temporary columns before returning results
            if '__key_tuple' in src_df.columns:
                src_df.drop(columns=['__key_tuple'], inplace=True)
            if '__key_tuple' in tgt_df.columns:
                tgt_df.drop(columns=['__key_tuple'], inplace=True)
                
            df_cmp_result_data = pd.concat(all_diffs) if all_diffs else pd.DataFrame()
            
            # Count value mismatches (each real mismatch has 2 rows - source and target)
            value_mismatch_count = len(value_mismatch_df) // 2 if not value_mismatch_df.empty else 0
            
            # Prepare standardized result dictionary
            total_diff_count = len(in_src_not_in_tgt) + len(in_tgt_not_in_src) + len(value_mismatch_df)
            is_identical = total_diff_count == 0

            if '__key_tuple' in df_cmp_result_data.columns:
                df_cmp_result_data.drop(columns=['__key_tuple'], inplace=True)
            
            result_dict = {
                'in_src_df_not_in_tgt_df': in_src_not_in_tgt,
                'in_tgt_df_not_in_src_df': in_tgt_not_in_src,
                'value_mismatch_df': value_mismatch_df,
                'combined_diff_df': df_cmp_result_data,
                'total_src_rows': len(src_df),
                'total_tgt_rows': len(tgt_df),
                'rows_only_in_src': len(in_src_not_in_tgt),
                'rows_only_in_tgt': len(in_tgt_not_in_src),
                'value_mismatches': value_mismatch_count,
                'total_diff_count': total_diff_count,
                'is_identical': is_identical,
                'mismatch_details': value_mismatch_details
            }

            del src_df
            del tgt_df
            del src_keys_df
            del tgt_keys_df
            del in_src_not_in_tgt
            del in_tgt_not_in_src
            del value_mismatch_df

            if 'src_grouped' in locals():
                del src_grouped
            if 'tgt_grouped' in locals():
                del tgt_grouped
            
            gc.collect()
            return result_dict
        
        except Exception as e:
            self.log.error(f"Failed: Enhanced key-based comparison - {e}")
            self.log.error(traceback.format_exc())
            gc.collect()
            raise ValueError(f"Enhanced key-based comparison failed: {e}")
    
        
    def check_threshold(self, src_rec_count: Union[int, float], 
                        tgt_rec_count: Union[int, float], 
                        threshold_percentage: Union[float, str]) -> Tuple[bool, str]:
        """
        Check if source record count is within acceptable threshold of target count.
        
        Args:
            src_rec_count: Source record count
            tgt_rec_count: Target record count
            threshold_percentage: Acceptable difference as a percentage
            
        Returns:
            Tuple with (is_within_threshold, message)
        """
        try:
            # Convert inputs to float for calculation
            src_count = float(src_rec_count)
            tgt_count = float(tgt_rec_count)
            threshold = float(threshold_percentage)
            
            # Calculate the threshold value
            threshold_value = tgt_count * threshold
            
            # Determine if counts match exactly
            if src_count == tgt_count:
                return True, 'Record Count Match Exactly'
            
            # Determine if within threshold
            elif abs(src_count - tgt_count) <= abs(threshold_value):
                return True, 'Record Count Match, with acceptable threshold'
            
            else:
                return False, 'Record Count Mismatch'
                
        except Exception as e:
            self.log.error(f"Failed: Threshold checking - {e}")
            raise ValueError(f"Threshold check failed: {e}")
    
    def compare_column_structure(self, src_df: pd.DataFrame, 
                               tgt_df: pd.DataFrame) -> Tuple[bool, set, set]:
        """
        Compare column structure between source and target dataframes.
        
        Args:
            src_df: Source dataframe
            tgt_df: Target dataframe
            
        Returns:
            Tuple with (columns_match, source_only_columns, target_only_columns)
        """
        try:
            # Get column lists
            src_col_list = src_df.columns.tolist()
            tgt_col_list = tgt_df.columns.tolist()
            
            # Flag is True if column counts match
            flag = len(src_col_list) == len(tgt_col_list)
            
            # Find columns only in source
            non_match_a = set(src_col_list) - set(tgt_col_list)
            
            # Find columns only in target
            non_match_b = set(tgt_col_list) - set(src_col_list)
            
            return flag, non_match_a, non_match_b
            
        except Exception as e:
            self.log.error(f"Failed: get non match columns - {e}")
            raise ValueError(f"Column comparison failed: {e}")
        
    def check_duplicates(self,df: pd.DataFrame, key_columns: list = None) -> pd.DataFrame:
        """
        Check for duplicate rows in a DataFrame based on specified key columns.
        
        Args:
            df (pd.DataFrame): Input DataFrame
            key_columns (list, optional): List of column names to check for duplicates. 
                                        If None, all columns are used.
        
        Returns:
            pd.DataFrame: DataFrame containing only the duplicate rows
        """
        try:
            if key_columns is None:
                # Use all columns if no specific key columns are provided
                duplicates = df[df.duplicated(keep=False)]
                duplicate_check_columns = list(df.columns)
            else:
                # Validate that all key columns exist in the DataFrame
                actual_key_columns = find_key_columns(df,key_columns)
                if not actual_key_columns:
                    raise ValueError(f"Key columns not found in DataFrame: {actual_key_columns}")
                    
                # Check for duplicates based on the specified key columns
                duplicates = df[df.duplicated(subset=key_columns, keep=False)]
                duplicate_check_columns = key_columns
                
            if duplicates.empty:
                self.log.info(f"No duplicates found in the dataset based on columns: {duplicate_check_columns}")
            else:
                # Count unique duplicate groups
                duplicate_groups = len(duplicates.drop_duplicates(subset=duplicate_check_columns))
                total_duplicates = len(duplicates)
                
                self.log.error(
                    f"Found {total_duplicates} duplicate rows in {duplicate_groups} groups "
                    f"based on columns: {duplicate_check_columns}"
                )
                
                # Optional: Add a column to identify duplicate groups
                if key_columns is not None:
                    # Create a group identifier for each set of duplicates
                    duplicates = duplicates.copy()
                    duplicates['duplicate_group'] = duplicates.groupby(key_columns).ngroup()
                    
            return duplicates
            
        except Exception as e:
            self.log.error(f"Failed to check for duplicates: {e}")
            raise


# Helper function for get_column_letter
def get_column_letter(idx):
    """Convert a column index to a column letter (e.g., 1 -> 'A', 27 -> 'AA')"""
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


class DataDiff:
    """
    Handles analysis, categorization, and visualization of mismatches between datasets.
    Takes output from DataComparator and provides detailed analysis and Excel export.
    """

    def __init__(self, logger=None):
        self.log = custom_logger.customLogger()
        self.comparison_summary = {}
   
    def analyze_and_categorize_mismatches(self, mismatch_df, key_columns, output_file=None, 
                                     max_value_mismatches=1000, max_source_only=250, 
                                     max_target_only=250, cmp_summary_df=None):
        """
        Analyzes and categorizes mismatches between source and target dataframes.
        
        Parameters:
            mismatch_df: DataFrame with comparison results
            key_columns: List of columns that uniquely identify records
            output_file: Path to save Excel output (optional)
            max_value_mismatches: Maximum number of value mismatch records to include
            max_source_only: Maximum number of source-only records to include
            max_target_only: Maximum number of target-only records to include
            cmp_summary_df: Optional comparison summary dataframe
            
        Returns:
            tuple: (result_dataframe, summary_dictionary)
        """
        try:
            self.log.info(f"Processing mismatch analysis with {len(mismatch_df)} rows")
            
            # Log mismatch_df contents for debugging
            if not mismatch_df.empty:
                #self.log.info(f"Mismatch DF columns: {list(mismatch_df.columns)}")
                
                # Check for expected columns
                if '__mismatch_type' in mismatch_df.columns:
                    mismatch_types = mismatch_df['__mismatch_type'].value_counts().to_dict()
                    self.log.info(f"Mismatch types in input: {mismatch_types}")
                    
                # IMPORTANT: Check for empty DataFrame with columns
                if len(mismatch_df) == 0:
                    self.log.warning("Empty mismatch DataFrame received")
            else:
                self.log.warning("Completely empty mismatch DataFrame received")
            
            # Always proceed even with empty DataFrame, as it may still have target-only rows in summary
            # Make a copy to avoid modifying the original
            df = mismatch_df.copy()
            
            # Step 1: Identify the source column
            source_col = self._find_source_column(df)
            if not source_col:
                self.log.error("Source column not found")
                # Create a fallback source column if missing
                if len(df) > 0:
                    df['__source'] = 'Unknown'
                    source_col = '__source'
                else:
                    # Create minimal info for reporting
                    if cmp_summary_df is not None:
                        summary_dict = dict(zip(cmp_summary_df['Metric'], cmp_summary_df['Value']))
                        target_only = summary_dict.get('Rows Only in Target', 0)
                        source_only = summary_dict.get('Rows Only in Source', 0)
                        
                        # If there are only target-only or source-only rows and no rows in df, create placeholder rows
                        if target_only > 0 or source_only > 0:
                            self.log.info(f"Summary reports {target_only} target-only rows and {source_only} source-only rows, but DataFrame is empty. Creating placeholder.")
                            # Create minimal placeholder for rows
                            cols = ['__source', '__mismatch_type'] + (key_columns if key_columns else [])
                            df = pd.DataFrame([{col: "PLACEHOLDER" for col in cols}])
                            if target_only > 0:
                                df['__source'] = 'Target'
                                df['__mismatch_type'] = 'Target-only-row'
                            else:
                                df['__source'] = 'Source'
                                df['__mismatch_type'] = 'Source-only-row'
                            source_col = '__source'
                        else:
                            return None, {"error": "Source column not found and dataframe is empty"}
                    else:
                        return None, {"error": "Source column not found and dataframe is empty"}
            
            # Step 2: Find actual key columns (case-insensitive matching)
            actual_key_columns = find_key_columns(df, key_columns, self.log)
            if not actual_key_columns:
                self.log.error(f"No key columns were found among: {key_columns}")
                # If there's summary data, we can still create a minimal result
                if cmp_summary_df is not None:
                    summary_dict = dict(zip(cmp_summary_df['Metric'], cmp_summary_df['Value']))
                    target_only = summary_dict.get('Rows Only in Target', 0)
                    source_only = summary_dict.get('Rows Only in Source', 0)
                    
                    if target_only > 0 or source_only > 0:
                        self.log.info(f"Using minimal columns since no key columns found but have differences")
                        # Create minimal placeholder
                        actual_key_columns = ['placeholder_key']
                        df['placeholder_key'] = 'key_value'
                    else:
                        return None, {"error": "No key columns were found"}
                else:
                    return None, {"error": "No key columns were found"}
            
            self.log.info(f"Using key columns: {actual_key_columns}")
            
            # Step 3: Separate source and target rows
            source_rows, target_rows = self._separate_source_target_rows(df, source_col)
            if source_rows is None and target_rows is None:
                self.log.error("Error separating source and target rows")
                return None, {"error": "Error separating source and target rows"}
                
            # Handle case where one of the dataframes might be empty but the other isn't
            if source_rows is None:
                #self.log.info("No source rows found, creating empty DataFrame with target columns")
                source_rows = pd.DataFrame(columns=target_rows.columns)
            
            if target_rows is None:
                #self.log.info("No target rows found, creating empty DataFrame with source columns")
                target_rows = pd.DataFrame(columns=source_rows.columns)
            
            #self.log.info(f"Found {len(source_rows)} source rows and {len(target_rows)} target rows")
            
            # Double check if any rows have __mismatch_type values
            if '__mismatch_type' in df.columns:
                if df['__mismatch_type'].notna().sum() == 0:
                    self.log.warning("__mismatch_type column exists but all values are null")
                    # Try to create mismatch types based on source values
                    if source_col in df.columns:
                        df.loc[df[source_col] == 'Source', '__mismatch_type'] = 'Source-only-row'
                        df.loc[df[source_col] == 'Target', '__mismatch_type'] = 'Target-only-row'
                        # Update source and target dataframes
                        source_rows = df[df[source_col] == 'Source'].copy()
                        target_rows = df[df[source_col] == 'Target'].copy()
            
            # Step 4: Create normalized composite keys for matching records
            try:
                if len(source_rows) > 0:
                    source_rows = create_composite_keys(source_rows, actual_key_columns, self.log)
                if len(target_rows) > 0:
                    target_rows = create_composite_keys(target_rows, actual_key_columns, self.log)
            except Exception as e:
                self.log.error(f"Error creating composite keys: {str(e)}")
                # Continue despite the error - we'll handle below
            
            # Verify keys were created
            if len(source_rows) > 0 and '__key' not in source_rows.columns:
                self.log.warning("Failed to create keys, implementing fallback approach")
                # Manual key creation as fallback
                def create_manual_key(row, columns):
                    return tuple(str(row[col]).lower() for col in columns if col in row)
                
                source_rows['__key'] = source_rows.apply(
                    lambda row: create_manual_key(row, actual_key_columns), axis=1
                )
                
            if len(target_rows) > 0 and '__key' not in target_rows.columns:
                self.log.warning("Failed to create keys in target, implementing fallback approach")
                # Manual key creation as fallback
                def create_manual_key(row, columns):
                    return tuple(str(row[col]).lower() for col in columns if col in row)
                    
                target_rows['__key'] = target_rows.apply(
                    lambda row: create_manual_key(row, actual_key_columns), axis=1
                )
            
            # Verify each dataframe has the __key column if not empty
            if len(source_rows) > 0 and '__key' not in source_rows.columns:
                self.log.error("Failed to create __key column in source_rows")
                source_rows['__key'] = range(len(source_rows))  # Fallback to row numbers
                
            if len(target_rows) > 0 and '__key' not in target_rows.columns:
                self.log.error("Failed to create __key column in target_rows")
                target_rows['__key'] = range(len(target_rows))  # Fallback to row numbers
            
            # Step 5: Process mismatches - USE THE UPDATED METHOD
            self.log.info("Processing mismatches...")
            result_df, mismatches = self._process_mismatches(
                source_rows, target_rows, actual_key_columns, source_col,
                max_value_mismatches, max_source_only, max_target_only
            )
            
            # CRITICAL: Check for cases where we only have one type of mismatch
            if result_df.empty:
                # Check for source-only rows
                if len(source_rows) > 0:
                    self.log.warning("Result dataframe is empty but source rows were found!")
                    # Verify source-only keys were identified
                    source_only_keys = mismatches.get('source_only_keys', set())
                    #self.log.info(f"Source-only keys found: {len(source_only_keys)}")
                    
                    if source_only_keys:
                        # Manually add source-only rows
                        source_only_rows = []
                        for key in list(source_only_keys)[:max_source_only]:
                            src_matches = source_rows[source_rows['__key'] == key]
                            if not src_matches.empty:
                                row_copy = src_matches.iloc[0].copy()
                                row_copy['__mismatch_type'] = 'Source-only-row'
                                source_only_rows.append(row_copy)
                        
                        if source_only_rows:
                            #self.log.info(f"Manually adding {len(source_only_rows)} source-only rows")
                            result_df = pd.DataFrame(source_only_rows)
                
                # Check for target-only rows if still empty
                if result_df.empty and len(target_rows) > 0:
                    self.log.warning("Result dataframe is empty but target rows were found!")
                    # Verify target-only keys were identified
                    target_only_keys = mismatches.get('target_only_keys', set())
                    #self.log.info(f"Target-only keys found: {len(target_only_keys)}")
                    
                    if target_only_keys:
                        # Manually add target-only rows
                        target_only_rows = []
                        for key in list(target_only_keys)[:max_target_only]:
                            tgt_matches = target_rows[target_rows['__key'] == key]
                            if not tgt_matches.empty:
                                row_copy = tgt_matches.iloc[0].copy()
                                row_copy['__mismatch_type'] = 'Target-only-row'
                                target_only_rows.append(row_copy)
                        
                        if target_only_rows:
                            #self.log.info(f"Manually adding {len(target_only_rows)} target-only rows")
                            result_df = pd.DataFrame(target_only_rows)
            
            # Check if result_df is still empty but we expect mismatches from summary data
            if result_df.empty and cmp_summary_df is not None:
                summary_dict = dict(zip(cmp_summary_df['Metric'], cmp_summary_df['Value']))
                source_only_count = summary_dict.get('Rows Only in Source', 0)
                target_only_count = summary_dict.get('Rows Only in Target', 0)
                
                # Check for source-only rows in summary
                if source_only_count > 0 and len(source_rows) > 0:
                    self.log.warning(f"Result dataframe is empty but expect {source_only_count} source-only rows - using raw source rows")
                    result_df = source_rows.copy()
                    if '__mismatch_type' not in result_df.columns:
                        result_df['__mismatch_type'] = 'Source-only-row'
                
                # Check for target-only rows in summary if still empty
                elif target_only_count > 0 and len(target_rows) > 0:
                    self.log.warning(f"Result dataframe is empty but expect {target_only_count} target-only rows - using raw target rows")
                    result_df = target_rows.copy()
                    if '__mismatch_type' not in result_df.columns:
                        result_df['__mismatch_type'] = 'Target-only-row'
                
                # If we still have no result but have input data, use it as fallback
                elif len(source_rows) > 0 or len(target_rows) > 0:
                    self.log.warning("Result dataframe is empty but input had data - using raw input as fallback")
                    # Use the original mismatch_df as fallback
                    result_df = mismatch_df.copy()
            
            # Log the counts from processing
            #self.log.info(f"Mismatch processing results:")
            #self.log.info(f"  Value mismatch count: {mismatches['value_mismatch_count']}")
            #self.log.info(f"  Source-only count: {mismatches['source_only_count']}")
            #self.log.info(f"  Target-only count: {mismatches['target_only_count']}")
            #self.log.info(f"  Final result_df size: {len(result_df)} rows")
            
            # Step 6: Generate summary statistics
            summary_stats = self._generate_summary_stats(
                source_rows, target_rows, result_df, mismatch_df, cmp_summary_df
            )
            
            # Step 7: Export to Excel if requested
            if output_file:
                mismatch_info = {
                    'mismatches': mismatches['mismatch_details'],
                    'actual_key_columns': actual_key_columns,
                    'source_col': source_col,
                    'mismatch_type_col': '__mismatch_type'
                }
                self._export_to_excel(result_df, mismatch_info, output_file)
            
            # Step 8: Clean up and return results
            if not result_df.empty:
                # Remove internal columns
                columns_to_remove = ['__key', '__key_tuple']
                for col in columns_to_remove:
                    if col in result_df.columns:
                        result_df = result_df.drop(columns=[col])
                
            self.log.info(f"Analysis complete. {len(result_df)} records in final output.")
            return result_df, summary_stats
            
        except Exception as e:
            self.log.error(f"Unexpected error in analyze_and_categorize_mismatches: {str(e)}")
            self.log.error(traceback.format_exc())
            return None, {"error": f"Unexpected error: {str(e)}"}

    def _find_source_column(self, df):
       """Find the source column in the dataframe"""
       try:
           source_col = [col for col in df.columns if col.lower() in ('__source', '_source')]
           if not source_col:
               self.log.error("Error: Source column ('__source' or '_source') not found in the dataframe")
               return None
           return source_col[0]
       except Exception as e:
           self.log.error(f"Error finding source column: {str(e)}")
           return None

    def _separate_source_target_rows(self, df, source_col):
        """Separate dataframe into source and target rows"""
        try:
            source_rows = df[df[source_col] == 'Source'].copy()
            target_rows = df[df[source_col] == 'Target'].copy()
            
            # If standard values not found, try to use the first two unique values
            unique_values = df[source_col].unique()
            #self.log.info(f"Values in source column: {unique_values}")
            
            # SPECIAL CASE: If only Source found but no Target
            if len(source_rows) > 0 and len(target_rows) == 0:
                #self.log.info(f"Found {len(source_rows)} Source rows but no Target rows - handling source-only case")
                # Create an empty target dataframe with same columns as source
                target_rows = pd.DataFrame(columns=source_rows.columns)
                return source_rows, target_rows
                
            # SPECIAL CASE: If only Target found but no Source
            if len(source_rows) == 0 and len(target_rows) > 0:
                #self.log.info(f"Found {len(target_rows)} Target rows but no Source rows - handling target-only case")
                # Create an empty source dataframe with same columns as target
                source_rows = pd.DataFrame(columns=target_rows.columns)
                return source_rows, target_rows
            
            # Handle case where the standard values 'Source' and 'Target' are not found
            if len(source_rows) == 0 or len(target_rows) == 0:
                #self.log.warning(f"Warning: No standard 'Source'/'Target' values found. Available values: {unique_values}")
                if len(unique_values) >= 2:
                    source_rows = df[df[source_col] == unique_values[0]].copy()
                    target_rows = df[df[source_col] == unique_values[1]].copy()
                elif len(unique_values) == 1:
                    # Only one value found, assume it's source and create empty target
                    #self.log.info(f"Only one value found ({unique_values[0]}), treating as Source and creating empty Target")
                    source_rows = df[df[source_col] == unique_values[0]].copy()
                    target_rows = pd.DataFrame(columns=source_rows.columns)
                else:
                    self.log.error(f"Not enough unique values in source column. Found: {unique_values}")
                    return None, None
                    
            return source_rows, target_rows
        except Exception as e:
            self.log.error(f"Error separating source and target rows: {str(e)}")
            self.log.error(traceback.format_exc())
            return None, None

    def _process_mismatches(self, source_rows, target_rows, key_columns, source_col,
                          max_value_mismatches, max_source_only, max_target_only):
        """Process and categorize mismatches between source and target"""
        try:
            # Get unique keys 
            source_keys = set(source_rows['__key']) if not source_rows.empty else set()
            target_keys = set(target_rows['__key']) if not target_rows.empty else set()
            
            # Find different types of keys
            common_keys = source_keys.intersection(target_keys)
            source_only_keys = source_keys - target_keys
            target_only_keys = target_keys - source_keys
            
            #self.log.info(f"Keys analysis: {len(common_keys)} common key combinations, " +
                        #f"{len(source_only_keys)} source-only combinations, " +
                        #f"{len(target_only_keys)} target-only combinations")
            
            # Identify value mismatches in common keys
            value_mismatch_keys = []
            mismatch_details = {}
            
            # Special columns to exclude from comparison
            special_cols = [source_col, '__key', '__mismatch_type']
            non_key_cols = [col for col in source_rows.columns 
                         if col not in key_columns and col not in special_cols]
            
            # Identify value mismatches (limited to max_value_mismatches)
            mismatch_count = 0
            for key in common_keys:
                if mismatch_count >= max_value_mismatches:
                    break
                
                src_rows = source_rows[source_rows['__key'] == key]
                tgt_rows = target_rows[target_rows['__key'] == key]
                
                if src_rows.empty or tgt_rows.empty:
                    continue
                
                # Compare all rows for this key
                key_mismatches = []
                
                for _, src_row in src_rows.iterrows():
                    for _, tgt_row in tgt_rows.iterrows():
                        # Compare non-key fields to find value mismatches
                        col_mismatches = []
                        
                        for col in non_key_cols:
                            if col in src_row and col in tgt_row:
                                src_val = src_row[col]
                                tgt_val = tgt_row[col]
                                
                                if not compare_values(src_val, tgt_val):
                                    col_mismatches.append(col)
                        
                        if col_mismatches:
                            # Keep track of which specific columns had mismatches
                            key_mismatches.extend(col_mismatches)
                            
                            # If this is a new mismatch key, add it to our list
                            if key not in value_mismatch_keys:
                                value_mismatch_keys.append(key)
                                
                            # Store a lookup key for highlighting
                            lookup_key = tuple(src_row[col] for col in key_columns)
                            if lookup_key not in mismatch_details:
                                mismatch_details[lookup_key] = col_mismatches
                            else:
                                # If we already have this key, add any new mismatched columns
                                mismatch_details[lookup_key].extend(
                                    [col for col in col_mismatches 
                                     if col not in mismatch_details[lookup_key]]
                                )
                
                if key_mismatches:
                    mismatch_count += 1
            
            # Build result dataframe
            filtered_rows = []
            
            # Add ALL rows for value mismatches
            value_mismatch_count = 0
            for key in value_mismatch_keys:
                src_matches = source_rows[source_rows['__key'] == key].copy()
                tgt_matches = target_rows[target_rows['__key'] == key].copy()
                
                for _, row in src_matches.iterrows():
                    row_copy = row.copy()
                    row_copy['__mismatch_type'] = 'Value-Mismatch'
                    filtered_rows.append(row_copy)
                    value_mismatch_count += 1
                
                for _, row in tgt_matches.iterrows():
                    row_copy = row.copy()
                    row_copy['__mismatch_type'] = 'Value-Mismatch'
                    filtered_rows.append(row_copy)
                    value_mismatch_count += 1
            
            # Process source-only rows, up to max_source_only
            source_only_count = 0
            #self.log.info(f"Processing {len(source_only_keys)} source-only keys (max {max_source_only})")
            
            # Make sure we process source-only rows even if they're the only mismatch type
            processed_source_keys = list(source_only_keys)[:max_source_only]
            #self.log.info(f"Will process {len(processed_source_keys)} source-only keys")
            
            for key in processed_source_keys:
                src_matches = source_rows[source_rows['__key'] == key]
                if not src_matches.empty:
                    row_copy = src_matches.iloc[0].copy()
                    row_copy['__mismatch_type'] = 'Source-only-row'
                    filtered_rows.append(row_copy)
                    source_only_count += 1
                    
            #self.log.info(f"Added {source_only_count} source-only rows to filtered_rows")
            
            # Process target-only rows, up to max_target_only
            target_only_count = 0
            #self.log.info(f"Processing {len(target_only_keys)} target-only keys (max {max_target_only})")
            
            # Make sure we process target-only rows even if they're the only mismatch type
            processed_target_keys = list(target_only_keys)[:max_target_only]
            #self.log.info(f"Will process {len(processed_target_keys)} target-only keys")
            
            for key in processed_target_keys:
                tgt_matches = target_rows[target_rows['__key'] == key]
                if not tgt_matches.empty:
                    row_copy = tgt_matches.iloc[0].copy()
                    row_copy['__mismatch_type'] = 'Target-only-row'
                    filtered_rows.append(row_copy)
                    target_only_count += 1
                    
            #self.log.info(f"Added {target_only_count} target-only rows to filtered_rows")
            
            # Create result dataframe - MAKE SURE WE HANDLE EMPTY CASE PROPERLY
            # Even if we only have target-only rows, we should create a valid dataframe
            if filtered_rows:
                result_df = pd.DataFrame(filtered_rows)
                #self.log.info(f"Created result_df with {len(result_df)} rows")
            else:
                self.log.warning("No rows in filtered_rows, creating empty DataFrame")
                # Create empty DataFrame with proper columns
                if not target_rows.empty:
                    result_df = pd.DataFrame(columns=target_rows.columns.tolist() + ['__mismatch_type'])
                elif not source_rows.empty:
                    result_df = pd.DataFrame(columns=source_rows.columns.tolist() + ['__mismatch_type'])
                else:
                    result_df = pd.DataFrame()
                    
            # Sort by key columns, if we have rows and key columns
            if not result_df.empty and all(col in result_df.columns for col in key_columns):
                result_df = result_df.sort_values(by=key_columns)
                
            # CRITICAL: Log the actual size and contents of result_df for debugging
            #self.log.info(f"Final result_df has {len(result_df)} rows")
            if not result_df.empty and '__mismatch_type' in result_df.columns:
                mismatch_counts = result_df['__mismatch_type'].value_counts().to_dict()
                #self.log.info(f"Mismatch types in result_df: {mismatch_counts}")
            
            # Make sure we account for the case where either source_only_count or target_only_count is the only non-zero count
            total_diff_count = value_mismatch_count // 2 + source_only_count + target_only_count
            #self.log.info(f"Total diff count: {total_diff_count} (VM: {value_mismatch_count//2}, SO: {source_only_count}, TO: {target_only_count})")
            
            return result_df, {
                'value_mismatch_keys': value_mismatch_keys,
                'source_only_keys': source_only_keys,
                'target_only_keys': target_only_keys,
                'mismatch_details': mismatch_details,
                'value_mismatch_count': value_mismatch_count // 2,  # Divide by 2 for pairs
                'source_only_count': source_only_count,
                'target_only_count': target_only_count
            }
        except Exception as e:
            self.log.error(f"Error processing mismatches: {str(e)}")
            self.log.error(traceback.format_exc())
            return pd.DataFrame(), {
                'value_mismatch_keys': [],
                'source_only_keys': [],
                'target_only_keys': [],
                'mismatch_details': {},
                'value_mismatch_count': 0,
                'source_only_count': 0,
                'target_only_count': 0
            }
    
    def _generate_summary_stats(self, source_rows, target_rows, result_df, mismatch_df, cmp_summary_df=None):
        """Generate summary statistics for the comparison"""
        try:
            # Count mismatch types more accurately
            if not result_df.empty and '__mismatch_type' in result_df.columns:
                # Get the raw counts from the result dataframe
                mismatch_counts = result_df['__mismatch_type'].value_counts().to_dict()
                
                # Calculate value mismatches (each value mismatch has a source and target row)
                value_mismatch_count = mismatch_counts.get('Value-Mismatch', 0) // 2  # Divide by 2 (source & target rows)
                source_only_count = mismatch_counts.get('Source-only-row', 0)
                target_only_count = mismatch_counts.get('Target-only-row', 0)
                
                # Log the counts to help with debugging
                #self.log.info(f"Mismatch counts from result_df: {mismatch_counts}")
                #self.log.info(f"Calculated value mismatch count: {value_mismatch_count}")
            else:
                value_mismatch_count = 0
                source_only_count = 0
                target_only_count = 0
            
            # Get total counts
            total_source_count = len(source_rows)
            total_target_count = len(target_rows)
            
            # Calculate the total differences
            total_differences = value_mismatch_count + source_only_count + target_only_count
            
            # Use provided summary if available, but ensure value mismatches are properly counted
            if cmp_summary_df is not None:
                try:
                    summary_dict = dict(zip(cmp_summary_df['Metric'], cmp_summary_df['Value']))
                    
                    # Use the provided summary data but override the value mismatch count
                    # if we have a non-zero value calculated above
                    provided_value_mismatches = summary_dict.get('Value Mismatches', 0)
                    final_value_mismatches = value_mismatch_count if value_mismatch_count > 0 else provided_value_mismatches
                    
                    # Get source and target only counts from summary if our local count is zero
                    provided_source_only = summary_dict.get('Rows Only in Source', 0)
                    provided_target_only = summary_dict.get('Rows Only in Target', 0)
                    
                    final_source_only = source_only_count if source_only_count > 0 else provided_source_only
                    final_target_only = target_only_count if target_only_count > 0 else provided_target_only
                    
                    # Recalculate total differences based on our potentially updated counts
                    final_total_diff = final_value_mismatches + final_source_only + final_target_only
                    
                    self.comparison_summary = {
                        'Total Source Rows': summary_dict.get('Total Source Rows', total_source_count),
                        'Total Target Rows': summary_dict.get('Total Target Rows', total_target_count),
                        'Comparison Result': 'Differences found' if final_total_diff > 0 else 'Identical',
                        'Rows Only in Source': final_source_only,
                        'Rows Only in Target': final_target_only,
                        'Value Mismatches': final_value_mismatches,
                        'Total Differences': final_total_diff,
                        'Value Mismatches in Output': value_mismatch_count,
                        'Source-only in Output': source_only_count,
                        'Target-only in Output': target_only_count
                    }
                    
                    #self.log.info(f"Summary from cmp_summary_df: {summary_dict}")
                    #self.log.info(f"Final comparison summary: {self.comparison_summary}")
                    
                except Exception as e:
                    self.log.error(f"Error using provided comparison summary: {str(e)}")
                    # Fall back to calculated summary
                    raise Exception("Falling back to calculated summary")
            else:
                # Calculate from the analyzed data
                self.comparison_summary = {
                    'Total Source Rows': total_source_count,
                    'Total Target Rows': total_target_count,
                    'Comparison Result': 'Differences found' if total_differences > 0 else 'Identical',
                    'Rows Only in Source': source_only_count,
                    'Rows Only in Target': target_only_count,
                    'Value Mismatches': value_mismatch_count,
                    'Total Differences': total_differences,
                    'Value Mismatches in Output': value_mismatch_count,
                    'Source-only in Output': source_only_count,
                    'Target-only in Output': target_only_count
                }
                
                self.log.info(f"Calculated comparison summary: {self.comparison_summary}")
            
            # Return summary for the caller
            return {
                'total_records': len(mismatch_df),
                'rows_only_in_src': source_only_count,
                'rows_only_in_tgt': target_only_count,
                'value_mismatches': value_mismatch_count,
                'records_in_output': len(result_df),
                'total_differences': total_differences
            }
        except Exception as e:
            self.log.error(f"Error generating summary statistics: {str(e)}")
            return {'error': f"Error creating summary: {str(e)}"}
        
    def _export_to_excel(self, result_df, mismatch_info, output_file):
        """Export the analyzed data to Excel with highlighting"""
        try:
                # Ensure output directory exists
                output_dir = os.path.dirname(output_file)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                    
                # Extract info for highlighting
                mismatches = mismatch_info['mismatches']
                actual_key_columns = mismatch_info['actual_key_columns']
                source_col = mismatch_info.get('source_col', '__source')
                mismatch_type_col = mismatch_info.get('mismatch_type_col', '__mismatch_type')
                
                # Log the mismatch details and result_df for debugging
                #self.log.info(f"Export to Excel - Result dataframe has {len(result_df)} rows")
                #self.log.info(f"Export to Excel - Mismatch details has {len(mismatches)} entries")
                
                # Check if result_df is empty but we have mismatches in the summary
                if len(result_df) == 0 and self.comparison_summary.get('Total Differences', 0) > 0:
                    self.log.warning("Result dataframe is empty but mismatches were detected in summary!")
                    # Create a basic info dataframe to indicate the issue
                    result_df = pd.DataFrame([
                        {"Warning": "Mismatch data exists but was not properly transferred to output",
                        "Details": f"Total differences detected: {self.comparison_summary.get('Total Differences', 'Unknown')}",
                        "Suggestion": "Check log files for more information"}
                    ])
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # Prepare output dataframe
                    output_df = result_df.copy()
                    # Remove internal columns that shouldn't be in the output
                    columns_to_remove = ['__key', '__key_tuple']
                    for col in columns_to_remove:
                        if col in output_df.columns:
                            output_df = output_df.drop(columns=[col])
                    
                    # Handle NaN values to prevent Excel errors
                    output_df = output_df.fillna('')
                    
                    # Handle empty dataframe case with improved error message
                    if len(output_df) == 0:
                        if self.comparison_summary.get('Total Differences', 0) > 0:
                            # We detected differences but they didn't make it to the output
                            error_msg = "Error: Differences were detected but not properly processed"
                            cause_msg = f"Summary shows {self.comparison_summary.get('Total Differences')} differences"
                        else:
                            # Truly no differences
                            error_msg = "No mismatch data found"
                            cause_msg = "No differences detected or filtering issue"
                        
                        output_df = pd.DataFrame([{"Error": error_msg, 
                                                "Possible cause": cause_msg}])
                    
                    # Write data to Excel
                    output_df.to_excel(writer, sheet_name='Mismatch Analysis', index=False)
                    
                    # Create summary sheet
                    self._create_summary_sheet(writer)
                    
                    # Apply highlighting to cells (only if we have actual data)
                    if len(result_df) > 0 and '__mismatch_type' in result_df.columns:
                        self._highlight_cells(writer, output_df, mismatch_info)
                    
                self.log.info(f"Successfully wrote data mismatch analysis results to {output_file}")
        except Exception as e:
                self.log.error(f"Unexpected error in export_to_excel: {str(e)}")
                self.log.error(traceback.format_exc())
                
                # Try a simpler export as fallback
                try:
                    output_df = result_df.copy()
                    # Remove internal columns in fallback method too
                    columns_to_remove = ['__key', '__key_tuple']
                    for col in columns_to_remove:
                        if col in output_df.columns:
                            output_df = output_df.drop(columns=[col])
                            
                    output_df.fillna('').to_excel(output_file, index=False)
                    self.log.info("Exported data using simplified method without formatting")
                except Exception as fallback_error:
                    self.log.error(f"Fallback export also failed: {str(fallback_error)}")

    def _create_summary_sheet(self, writer):
        """Create a summary sheet in the Excel workbook"""
        try:
            summary_rows = [
                {'Section': 'Dataset Comparison', 'Metric': 'Total Source Rows', 'Value': self.comparison_summary.get('Total Source Rows', 'N/A')},
                {'Section': 'Dataset Comparison', 'Metric': 'Total Target Rows', 'Value': self.comparison_summary.get('Total Target Rows', 'N/A')},
                {'Section': 'Dataset Comparison', 'Metric': 'Comparison Result', 'Value': self.comparison_summary.get('Comparison Result', 'N/A')},
                {'Section': 'Differences', 'Metric': 'Rows Only in Source', 'Value': self.comparison_summary.get('Rows Only in Source', 'N/A')},
                {'Section': 'Differences', 'Metric': 'Rows Only in Target', 'Value': self.comparison_summary.get('Rows Only in Target', 'N/A')},
                {'Section': 'Differences', 'Metric': 'Value Mismatches', 'Value': self.comparison_summary.get('Value Mismatches', 'N/A')},
                {'Section': 'Differences', 'Metric': 'Total Differences', 'Value': self.comparison_summary.get('Total Differences', 'N/A')},
                {'Section': 'Note', 'Metric': 'Sample Data', 
                    'Value': f'The Mismatch Analysis sheet contains up to {self.comparison_summary.get("Value Mismatches in Output", "N/A")} value mismatches, ' +
                            f'{self.comparison_summary.get("Source-only in Output", "N/A")} source-only rows, and ' +
                            f'{self.comparison_summary.get("Target-only in Output", "N/A")} target-only rows.'}
            ]
            
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_sheet = writer.sheets['Summary']
            
            # Set column widths
            for col in range(1, 4):
                col_letter = get_column_letter(col)
                summary_sheet.column_dimensions[col_letter].width = 25
            
            # Add background colors to sections
            for row in range(2, len(summary_rows) + 2):
                if row <= 4:  # Dataset Comparison
                    section_fill = PatternFill(start_color="E9F1FD", end_color="E9F1FD", fill_type="solid")  # Light blue
                elif row <= 8:  # Differences
                    section_fill = PatternFill(start_color="FFF8E8", end_color="FFF8E8", fill_type="solid")  # Light yellow
                else:  # Note
                    section_fill = PatternFill(start_color="EDFCED", end_color="EDFCED", fill_type="solid")  # Light green
                
                # Apply fill to row
                for col in range(1, 4):
                    cell = summary_sheet[f"{get_column_letter(col)}{row}"]
                    cell.fill = section_fill
                    
                    # Make section and metric headers bold
                    if col in [1, 2]:
                        cell.font = Font(bold=True)
        except Exception as e:
            self.log.error(f"Error creating summary sheet: {str(e)}")

    def _highlight_cells(self, writer, output_df, mismatch_info):
        """Apply highlighting to cells in the Excel worksheet"""
        try:
                # Get worksheet and define colors
                worksheet = writer.sheets['Mismatch Analysis']
                value_mismatch_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                source_only_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")     # Light red
                target_only_fill = PatternFill(start_color="CCEEFF", end_color="CCEEFF", fill_type="solid")     # Light blue
                
                # Extract info for highlighting
                mismatches = mismatch_info['mismatches']
                actual_key_columns = mismatch_info['actual_key_columns']
                mismatch_type_col = mismatch_info['mismatch_type_col']
                
                # Get column indices
                col_indices = {col_name: idx for idx, col_name in enumerate(output_df.columns, start=1)}
                
                # Log information about mismatches for debugging
                #self.log.info(f"Highlighting cells with {len(mismatches)} mismatch entries")
                #self.log.info(f"Mismatch keys: {list(mismatches.keys())[:5]}..." if mismatches else "No mismatches")
                
                # Track rows for highlighting
                rows_by_mismatch_type = {
                    'Value-Mismatch': [],
                    'Source-only-row': [],
                    'Target-only-row': []
                }
                
                # Group rows by mismatch type for more efficient processing
                for row_idx, (_, row) in enumerate(output_df.iterrows(), start=2):  # Excel rows start at 1, header at row 1
                    try:
                        mismatch_type = row[mismatch_type_col]
                        if mismatch_type in rows_by_mismatch_type:
                            # Store row index and key for this row
                            key_vals = tuple(row[k] for k in actual_key_columns)
                            rows_by_mismatch_type[mismatch_type].append((row_idx, key_vals, row))
                    except Exception as e:
                        self.log.warning(f"Error processing row {row_idx} for highlighting: {str(e)}")
                        continue
                
                # Log counts for verification
                #for mtype, rows in rows_by_mismatch_type.items():
                    #self.log.info(f"Found {len(rows)} rows of type {mtype}")
                
                # Highlight value mismatch fields
                highlighted_cells = 0
                value_mismatch_rows = rows_by_mismatch_type['Value-Mismatch']
                
                # Collect keys with mismatches and their mismatched columns
                key_to_mismatched_cols = {}
                
                # First pass: build key to mismatched columns mapping
                for key, mismatch_detail in mismatches.items():
                    # Handle both tuple values and lists for mismatch details
                    if isinstance(mismatch_detail, list):
                        # If the detail is a list of columns directly
                        key_to_mismatched_cols[key] = mismatch_detail
                    elif isinstance(mismatch_detail, dict) and 'cols' in mismatch_detail:
                        # If using the new structured format with 'cols' key
                        key_to_mismatched_cols[key] = mismatch_detail['cols']
                    else:
                        self.log.warning(f"Unexpected mismatch detail format: {mismatch_detail}")
                
                # If no mappings were found, try to infer from output_df
                if not key_to_mismatched_cols:
                    self.log.warning("No mismatch details found, attempting to infer from data")
                    # Group value mismatch rows by key
                    key_rows = {}
                    for row_idx, key_vals, row in value_mismatch_rows:
                        if key_vals not in key_rows:
                            key_rows[key_vals] = []
                        key_rows[key_vals].append(row)
                    
                    # Find columns with different values for each key
                    for key_vals, rows in key_rows.items():
                        if len(rows) >= 2:  # Need at least a source and target row
                            source_rows = [r for r in rows if r['__source'] == 'Source']
                            target_rows = [r for r in rows if r['__source'] == 'Target']
                            
                            if source_rows and target_rows:
                                # Take first row from each side for simplicity
                                src_row = source_rows[0]
                                tgt_row = target_rows[0]
                                
                                # Find mismatched columns
                                mismatched_cols = []
                                for col in output_df.columns:
                                    if col not in actual_key_columns and col not in ['__source', '__mismatch_type', '__key']:
                                        if src_row[col] != tgt_row[col]:
                                            mismatched_cols.append(col)
                                
                                if mismatched_cols:
                                    key_to_mismatched_cols[key_vals] = mismatched_cols
                
                # Now highlight the cells
                for row_idx, key_vals, row in value_mismatch_rows:
                    # Look for this key in our mapping
                    mismatch_found = False
                    
                    # Try direct key match
                    if key_vals in key_to_mismatched_cols:
                        mismatched_cols = key_to_mismatched_cols[key_vals]
                        mismatch_found = True
                    else:
                        # Try a more flexible matching approach
                        for stored_key, cols in key_to_mismatched_cols.items():
                            # Match if all key components match (case-insensitive)
                            if len(stored_key) == len(key_vals) and all(
                                str(k1).lower() == str(k2).lower() 
                                for k1, k2 in zip(stored_key, key_vals)
                            ):
                                mismatched_cols = cols
                                mismatch_found = True
                                break
                    
                    if mismatch_found:
                        # Highlight each mismatched column
                        for col in mismatched_cols:
                            if col in col_indices:
                                col_letter = get_column_letter(col_indices[col])
                                cell = worksheet[f"{col_letter}{row_idx}"]
                                cell.fill = value_mismatch_fill
                                highlighted_cells += 1
                
                # Highlight source-only rows
                for row_idx, _, _ in rows_by_mismatch_type['Source-only-row']:
                    for col_idx in range(1, len(output_df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        worksheet[f"{col_letter}{row_idx}"].fill = source_only_fill
                
                # Highlight target-only rows
                for row_idx, _, _ in rows_by_mismatch_type['Target-only-row']:
                    for col_idx in range(1, len(output_df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        worksheet[f"{col_letter}{row_idx}"].fill = target_only_fill
                
                #self.log.info(f"Highlighted {highlighted_cells} cells with value mismatches")
                
                # Add legend
                legend_row = len(output_df) + 5
                worksheet[f"A{legend_row}"] = "Legend:"
                worksheet[f"A{legend_row+1}"] = "Yellow"
                worksheet[f"A{legend_row+2}"] = "Light Red"
                worksheet[f"A{legend_row+3}"] = "Light Blue"
                worksheet[f"B{legend_row+1}"] = "Mismatched values (highlighted cells)"
                worksheet[f"B{legend_row+2}"] = "Source-only rows (entire row)"
                worksheet[f"B{legend_row+3}"] = "Target-only rows (entire row)"
                
                # Apply legend colors
                worksheet[f"A{legend_row+1}"].fill = value_mismatch_fill
                worksheet[f"A{legend_row+2}"].fill = source_only_fill
                worksheet[f"A{legend_row+3}"].fill = target_only_fill
            
        except Exception as e:
                self.log.error(f"Error highlighting cells in worksheet: {str(e)}")
                self.log.error(traceback.format_exc())